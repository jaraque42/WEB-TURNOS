from datetime import date, datetime
from io import BytesIO, StringIO
from typing import List, Optional

import csv
import httpx
import re
import unicodedata
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.assignment import ShiftAssignment, AssignmentStatus
from app.models.employee import Employee
from app.models.shift import ShiftType
from app.schemas.assignment import ShiftAssignmentCreate
from app.services.auth import require_permission

router = APIRouter(tags=["Importaciones"])

_ALIAS_MAP = {
    "date": {"date", "fecha", "dia", "day"},
    "employee_id": {"employee_id", "empleado_id", "id_empleado", "idempleado"},
    "email": {"email", "correo", "mail", "e_mail", "employee_email"},
    "dni": {"dni", "documento", "document_number", "documentnumber", "cedula", "nif"},
    "shift_type_id": {"shift_type_id", "id_turno", "shift_id", "tipo_turno_id"},
    "shift_code": {"shift_code", "shift_type_code", "codigo_turno", "cod_turno", "turno_codigo"},
    "shift_name": {"shift_name", "shift_type", "turno", "nombre_turno", "tipo_turno"},
    "location": {"location", "ubicacion", "site", "sede", "base"},
    "notes": {"notes", "observaciones", "comentarios", "comentario", "nota"},
}


def _normalize_key(key: str) -> str:
    text = unicodedata.normalize("NFKD", str(key)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text.strip().lower())
    return re.sub(r"_+", "_", text).strip("_")


def _canonical_key(raw_key: str) -> str:
    nk = _normalize_key(raw_key)
    for canonical, aliases in _ALIAS_MAP.items():
        if nk in aliases:
            return canonical
    return nk


def _normalize_row(row: dict) -> dict:
    normalized: dict = {}
    for k, v in row.items():
        if k is None:
            continue
        ck = _canonical_key(k)
        val = v.strip() if isinstance(v, str) else v
        if ck not in normalized or normalized[ck] in (None, ""):
            normalized[ck] = val
    return normalized


def _parse_date(value: object) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = str(value).strip()
    if not raw:
        return None

    # ISO (2026-03-01 o 2026-03-01T08:00:00)
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        pass

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


async def _resolve_employee(db: AsyncSession, row: dict) -> Optional[Employee]:
    # Preferimos employee_id UUID si viene. Luego email o DNI.
    emp = None
    emp_id = (row.get("employee_id") or "").strip() or None
    email = (row.get("email") or "").strip() or None
    dni = (row.get("dni") or row.get("document_number") or "").strip() or None
    if emp_id:
        result = await db.execute(select(Employee).where(Employee.id == emp_id))
        emp = result.scalar_one_or_none()
    elif email:
        result = await db.execute(select(Employee).where(func.lower(Employee.email) == email.lower()))
        emp = result.scalar_one_or_none()
    elif dni:
        result = await db.execute(select(Employee).where(Employee.document_number == dni))
        emp = result.scalar_one_or_none()
    return emp


async def _resolve_shift_type(db: AsyncSession, row: dict) -> Optional[ShiftType]:
    # Permitimos shift_type_id, code o name
    st = None
    st_id = (row.get("shift_type_id") or "").strip() or None
    code = (row.get("shift_code") or row.get("shift_type_code") or row.get("codigo_turno") or "").strip() or None
    name = (row.get("shift_name") or row.get("shift_type") or row.get("turno") or "").strip() or None
    if st_id:
        result = await db.execute(select(ShiftType).where(ShiftType.id == st_id))
        st = result.scalar_one_or_none()
    elif code:
        result = await db.execute(select(ShiftType).where(func.lower(ShiftType.code) == code.lower()))
        st = result.scalar_one_or_none()
    elif name:
        result = await db.execute(select(ShiftType).where(func.lower(ShiftType.name) == name.lower()))
        st = result.scalar_one_or_none()
    return st


def _parse_csv(content: bytes) -> List[dict]:
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(StringIO(text), dialect=dialect)
    rows = []
    for r in reader:
        cleaned = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in r.items() if k is not None}
        rows.append(_normalize_row(cleaned))
    return rows


def _parse_xlsx(content: bytes) -> List[dict]:
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl requerido para Excel: {e}")
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        data = {}
        for i, cell in enumerate(row):
            val = cell.value
            if isinstance(val, datetime):
                val = val.date().isoformat()
            data[headers[i]] = str(val).strip() if isinstance(val, str) else val
        rows.append(_normalize_row(data))
    return rows


async def _ingest_rows_as_assignments(db: AsyncSession, rows: List[dict]) -> dict:
    ok, errors = 0, []
    for idx, r in enumerate(rows, start=2):  # header en fila 1
        if all(v in (None, "") for v in r.values()):
            continue
        # Campos esperados: date, employee_id|email|dni, shift_type_id|shift_code|shift_name, location(opc), notes(opc)
        raw_date = r.get("date")
        if raw_date in (None, ""):
            errors.append({"row": idx, "error": "Falta columna date/fecha"})
            continue
        d = _parse_date(raw_date)
        if not d:
            errors.append({"row": idx, "error": f"Fecha inválida: {raw_date}"})
            continue

        emp = await _resolve_employee(db, r)
        if not emp:
            errors.append({
                "row": idx,
                "error": (
                    "Empleado no encontrado "
                    f"(employee_id={r.get('employee_id')}, email={r.get('email')}, dni={r.get('dni') or r.get('document_number')})"
                ),
            })
            continue

        st = await _resolve_shift_type(db, r)
        if not st:
            errors.append({
                "row": idx,
                "error": (
                    "Tipo de turno no encontrado "
                    f"(shift_type_id={r.get('shift_type_id')}, shift_code={r.get('shift_code') or r.get('shift_type_code') or r.get('codigo_turno')}, "
                    f"shift_name={r.get('shift_name') or r.get('shift_type') or r.get('turno')})"
                ),
            })
            continue

        location = r.get("location") or r.get("ubicacion") or None
        notes = r.get("notes") or None

        # Upsert por restricción única (employee_id, date)
        result = await db.execute(
            select(ShiftAssignment).where(
                (ShiftAssignment.employee_id == emp.id) & (ShiftAssignment.date == d)
            )
        )
        existing = result.scalar_one_or_none()
        if existing:
            existing.shift_type_id = st.id
            existing.location = location
            existing.notes = notes
            existing.status = AssignmentStatus.ASSIGNED
        else:
            db.add(ShiftAssignment(
                date=d,
                employee_id=emp.id,
                shift_type_id=st.id,
                location=location,
                notes=notes,
                status=AssignmentStatus.ASSIGNED,
            ))
        ok += 1
    await db.commit()
    return {"rows_ok": ok, "rows_error": len(errors), "errors": errors}


@router.post("/imports/forecasts/file")
async def import_forecasts_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: object = Depends(require_permission("assignments:create")),
):
    content = await file.read()
    ext = (file.filename or "").lower()
    if ext.endswith(".csv"):
        rows = _parse_csv(content)
    elif ext.endswith(".xlsx") or ext.endswith(".xlsm"):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Use .csv o .xlsx")
    result = await _ingest_rows_as_assignments(db, rows)
    sample = rows[:10]
    return {"result": result, "sample": sample}


@router.post("/imports/forecasts/url")
async def import_forecasts_url(
    url: str,
    db: AsyncSession = Depends(get_db),
    _: object = Depends(require_permission("assignments:create")),
):
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url)
        if resp.status_code != 200:
            raise HTTPException(status_code=400, detail=f"No se pudo descargar: {resp.status_code}")
        content_type = resp.headers.get("content-type", "").lower()
        content = resp.content
        if "text/csv" in content_type or url.lower().endswith(".csv"):
            rows = _parse_csv(content)
        elif "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content_type or url.lower().endswith(".xlsx"):
            rows = _parse_xlsx(content)
        else:
            # Intentar CSV por defecto
            try:
                rows = _parse_csv(content)
            except Exception:
                raise HTTPException(status_code=400, detail="Contenido no soportado. Use CSV/XLSX")
    result = await _ingest_rows_as_assignments(db, rows)
    sample = rows[:10]
    return {"result": result, "sample": sample}


@router.get("/imports/forecasts/summary")
async def forecasts_summary(
    date_from: date = Query(...),
    date_to: date = Query(...),
    group_by: str = Query("location", pattern="^(location|employee)$"),
    db: AsyncSession = Depends(get_db),
    _: object = Depends(require_permission("assignments:read")),
):
    if group_by not in ("location", "employee"):
        raise HTTPException(status_code=400, detail="group_by debe ser location|employee")

    if group_by == "location":
        q = (
            select(
                ShiftAssignment.date.label("d"),
                (ShiftAssignment.location).label("g"),
                func.count(ShiftAssignment.id).label("value"),
            )
            .where((ShiftAssignment.date >= date_from) & (ShiftAssignment.date <= date_to))
            .group_by(ShiftAssignment.date, ShiftAssignment.location)
            .order_by(ShiftAssignment.date)
        )
    else:
        # por empleado
        from app.models.employee import Employee
        q = (
            select(
                ShiftAssignment.date.label("d"),
                (Employee.full_name).label("g"),
                func.count(ShiftAssignment.id).label("value"),
            )
            .join(Employee, Employee.id == ShiftAssignment.employee_id)
            .where((ShiftAssignment.date >= date_from) & (ShiftAssignment.date <= date_to))
            .group_by(ShiftAssignment.date, Employee.full_name)
            .order_by(ShiftAssignment.date)
        )

    result = await db.execute(q)
    rows = result.all()
    # Estructura: { labels: [fechas], series: [{name, data: [valores por fecha]}] }
    labels = sorted({r.d for r in rows})
    groups = sorted({(r.g or "Sin ubicación") for r in rows})
    series_map = {g: [0] * len(labels) for g in groups}
    label_index = {d: i for i, d in enumerate(labels)}
    for r in rows:
        g = r.g or "Sin ubicación"
        i = label_index[r.d]
        series_map[g][i] = int(r.value or 0)

    series = [{"name": g, "data": vals} for g, vals in series_map.items()]
    return {"labels": [d.isoformat() for d in labels], "series": series}
